import openpyxl
from typing import Any

source_path = 'data/tax-return.xlsx'
file = openpyxl.load_workbook(source_path, read_only=True)


def get_excel(sheet_index: str):
    return file[sheet_index]


def get_amount_payment(sheet_index: int, max_row: int) -> int:
    file = openpyxl.load_workbook(source_path)
    sheet = file[sheet_index]
    sums = []
    AMOUNT = columns['paid']
    for i in range(1, max_row, 1):
        data = sheet.cell(i, AMOUNT).value
        if data is not None:
            sums.append(data)
    file.close()
    return sum(sums)


totals = [
    'hiroko-medical',   # 3, 14
    'hiroko-care',      # 3, 15
    'takashi-medical',
    'takashi-care'
    'medical-subtotal',
    'care-subtotal',
    'total'
]


def get_total_amount(select_index: int = 6):
    sheet = get_excel(sheet_index[4])

    def boolean(index: int, cell_x, cell_y):
        if index == select_index:
            return sheet.cell(cell_x, cell_y).value
        else:
            return None


class GetTotalAmount:
    def __init__(self, index: int):
        pass


def sentence(ret: int, index: int, max: int) -> int:
    path = 'data/tax-return.xlsx'
    file = openpyxl.load_workbook(path)
    sheet = file[index]
    payment_sum = []

    for i in range(1, max, 1):
        month = sheet.cell(i, columns['month']).value
        name = sheet.cell(i, columns['name']).value
        paid = sheet.cell(i, columns['paid']).value
        payment_sum.append(paid)

        if ret == 0:  # * amount payment
            return sum(payment_sum)
        elif ret == 1:  # * sentence
            ap = f'{month}月, {name}, {paid}円'
            payment_sum.append(ap)
            return payment_sum
        else:
            return None

    file.close()


rows = [
    33,
    28,
    55,
    55
]

columns = {
    'month': 1,
    'name': 2,
    'paid': 3
}

sheet_index = [
    'hiroko-med',
    'hiroko-care',
    'takashi-med',
    'takashi-care',
    'total'
]

load_sentences = [
    f'{index}: {get_amount_payment(index, row)}'
    for index, row in zip(sheet_index, rows)
] + [
    str(get_total_amount(select_index=0))
]

if __name__ == '__main__':
    for i in load_sentences:
        print(i)
